from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import PieChart, BarChart, Reference
from io import BytesIO
from django.conf import settings
import os


def generate_project_pdf(project, options=None):
    """Generate PDF report for a project"""
    if options is None:
        options = {}
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    story.append(Paragraph(f"Project Report: {project.name}", title_style))
    story.append(Spacer(1, 20))
    
    # Project Information
    story.append(Paragraph("Project Information", styles['Heading2']))
    
    project_data = [
        ['Project Name:', project.name],
        ['Type:', project.get_project_type_display()],
        ['Status:', project.get_status_display()],
        ['Location:', f"{project.address}, {project.city}, {project.postcode}"],
        ['Total Area:', f"{project.total_area} m²"],
        ['Floors:', str(project.floors)],
        ['Owner:', project.owner.full_name],
        ['Created:', project.created_at.strftime('%Y-%m-%d')],
    ]
    
    if project.start_date:
        project_data.append(['Start Date:', project.start_date.strftime('%Y-%m-%d')])
    if project.end_date:
        project_data.append(['End Date:', project.end_date.strftime('%Y-%m-%d')])
    
    project_table = Table(project_data, colWidths=[2*inch, 4*inch])
    project_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(project_table)
    story.append(Spacer(1, 20))
    
    # Estimates Summary
    estimates = project.estimates.all()
    if estimates:
        story.append(Paragraph("Cost Estimates", styles['Heading2']))
        
        estimate_data = [['Estimate Name', 'Status', 'Total Cost', 'Created']]
        for estimate in estimates:
            estimate_data.append([
                estimate.name,
                estimate.get_status_display(),
                f"£{estimate.total_cost:,.2f}",
                estimate.created_at.strftime('%Y-%m-%d')
            ])
        
        estimate_table = Table(estimate_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1*inch])
        estimate_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(estimate_table)
        story.append(Spacer(1, 20))
    
    # Collaborators
    collaborators = project.collaborators.all()
    if collaborators:
        story.append(Paragraph("Project Collaborators", styles['Heading2']))
        
        collab_data = [['Name', 'Email', 'Role']]
        for collaboration in project.projectcollaborator_set.all():
            collab_data.append([
                collaboration.user.full_name,
                collaboration.user.email,
                collaboration.get_role_display()
            ])
        
        collab_table = Table(collab_data, colWidths=[2*inch, 2.5*inch, 1.5*inch])
        collab_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(collab_table)
    
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def generate_estimate_pdf(estimate, options=None):
    """Generate PDF report for an estimate"""
    if options is None:
        options = {}
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=1
    )
    story.append(Paragraph(f"Cost Estimate: {estimate.name}", title_style))
    story.append(Spacer(1, 20))
    
    # Estimate Information
    story.append(Paragraph("Estimate Summary", styles['Heading2']))
    
    summary_data = [
        ['Project:', estimate.project.name],
        ['Estimate Name:', estimate.name],
        ['Status:', estimate.get_status_display()],
        ['Created By:', estimate.created_by.full_name],
        ['Created Date:', estimate.created_at.strftime('%Y-%m-%d')],
        ['', ''],
        ['Materials Cost:', f"£{estimate.materials_cost:,.2f}"],
        ['Labor Cost:', f"£{estimate.labor_cost:,.2f}"],
        ['Machinery Cost:', f"£{estimate.machinery_cost:,.2f}"],
        ['Overhead Cost:', f"£{estimate.overhead_cost:,.2f}"],
        ['Subtotal:', f"£{estimate.subtotal:,.2f}"],
        ['VAT ({:.1%}):'.format(estimate.vat_rate), f"£{estimate.vat_amount:,.2f}"],
        ['TOTAL COST:', f"£{estimate.total_cost:,.2f}"],
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 4*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        # Highlight total row
        ('BACKGROUND', (0, -1), (-1, -1), colors.yellow),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 30))
    
    # Materials Breakdown
    if estimate.material_items.exists():
        story.append(Paragraph("Materials Breakdown", styles['Heading2']))
        
        material_data = [['Material', 'Quantity', 'Unit Price', 'Total Cost', 'Supplier']]
        for item in estimate.material_items.all():
            material_data.append([
                item.material.name,
                f"{item.quantity} {item.material.unit}",
                f"£{item.unit_price:,.2f}",
                f"£{item.total_cost:,.2f}",
                item.supplier or 'TBD'
            ])
        
        material_table = Table(material_data, colWidths=[2*inch, 1*inch, 1*inch, 1*inch, 1*inch])
        material_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(material_table)
        story.append(Spacer(1, 20))
    
    # Machinery Breakdown
    if estimate.machinery_items.exists():
        story.append(Paragraph("Machinery Breakdown", styles['Heading2']))
        
        machinery_data = [['Equipment', 'Type', 'Duration', 'Unit Price', 'Total Cost']]
        for item in estimate.machinery_items.all():
            machinery_data.append([
                item.machinery.name,
                item.get_rental_type_display(),
                str(item.duration),
                f"£{item.unit_price:,.2f}",
                f"£{item.total_cost:,.2f}"
            ])
        
        machinery_table = Table(machinery_data, colWidths=[2*inch, 1*inch, 1*inch, 1*inch, 1*inch])
        machinery_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(machinery_table)
    
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def generate_estimate_excel(estimate, options=None):
    """Generate Excel report for an estimate"""
    if options is None:
        options = {}
    
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Summary Sheet
    ws_summary = wb.create_sheet("Summary")
    
    # Headers
    ws_summary['A1'] = "Cost Estimate Summary"
    ws_summary['A1'].font = Font(size=16, bold=True)
    
    # Estimate info
    ws_summary['A3'] = "Project:"
    ws_summary['B3'] = estimate.project.name
    ws_summary['A4'] = "Estimate:"
    ws_summary['B4'] = estimate.name
    ws_summary['A5'] = "Status:"
    ws_summary['B5'] = estimate.get_status_display()
    ws_summary['A6'] = "Created:"
    ws_summary['B6'] = estimate.created_at.strftime('%Y-%m-%d')
    
    # Cost breakdown
    ws_summary['A8'] = "Cost Breakdown"
    ws_summary['A8'].font = Font(bold=True)
    
    ws_summary['A9'] = "Materials:"
    ws_summary['B9'] = float(estimate.materials_cost)
    ws_summary['A10'] = "Labor:"
    ws_summary['B10'] = float(estimate.labor_cost)
    ws_summary['A11'] = "Machinery:"
    ws_summary['B11'] = float(estimate.machinery_cost)
    ws_summary['A12'] = "Overhead:"
    ws_summary['B12'] = float(estimate.overhead_cost)
    ws_summary['A13'] = "Subtotal:"
    ws_summary['B13'] = float(estimate.subtotal)
    ws_summary['A14'] = f"VAT ({estimate.vat_rate:.1%}):"
    ws_summary['B14'] = float(estimate.vat_amount)
    ws_summary['A15'] = "TOTAL:"
    ws_summary['B15'] = float(estimate.total_cost)
    ws_summary['A15'].font = Font(bold=True)
    ws_summary['B15'].font = Font(bold=True)
    
    # Format currency columns
    for row in range(9, 16):
        ws_summary[f'B{row}'].number_format = '£#,##0.00'
    
    # Materials Sheet
    if estimate.material_items.exists():
        ws_materials = wb.create_sheet("Materials")
        
        # Headers
        headers = ['Material', 'Category', 'Quantity', 'Unit', 'Unit Price', 'Waste Factor', 'Total Cost', 'Supplier']
        for col, header in enumerate(headers, 1):
            cell = ws_materials.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        for row, item in enumerate(estimate.material_items.all(), 2):
            ws_materials.cell(row=row, column=1, value=item.material.name)
            ws_materials.cell(row=row, column=2, value=item.material.category.name)
            ws_materials.cell(row=row, column=3, value=float(item.quantity))
            ws_materials.cell(row=row, column=4, value=item.material.unit)
            ws_materials.cell(row=row, column=5, value=float(item.unit_price))
            ws_materials.cell(row=row, column=6, value=float(item.waste_factor))
            ws_materials.cell(row=row, column=7, value=float(item.total_cost))
            ws_materials.cell(row=row, column=8, value=item.supplier or 'TBD')
            
            # Format currency
            ws_materials.cell(row=row, column=5).number_format = '£#,##0.00'
            ws_materials.cell(row=row, column=7).number_format = '£#,##0.00'
            # Format percentage
            ws_materials.cell(row=row, column=6).number_format = '0.0%'
    
    # Machinery Sheet
    if estimate.machinery_items.exists():
        ws_machinery = wb.create_sheet("Machinery")
        
        # Headers
        headers = ['Equipment', 'Category', 'Rental Type', 'Duration', 'Unit Price', 'Transport', 'Setup', 'Total Cost', 'Supplier']
        for col, header in enumerate(headers, 1):
            cell = ws_machinery.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        for row, item in enumerate(estimate.machinery_items.all(), 2):
            ws_machinery.cell(row=row, column=1, value=item.machinery.name)
            ws_machinery.cell(row=row, column=2, value=item.machinery.category.name)
            ws_machinery.cell(row=row, column=3, value=item.get_rental_type_display())
            ws_machinery.cell(row=row, column=4, value=float(item.duration))
            ws_machinery.cell(row=row, column=5, value=float(item.unit_price))
            ws_machinery.cell(row=row, column=6, value=float(item.transport_cost))
            ws_machinery.cell(row=row, column=7, value=float(item.setup_cost))
            ws_machinery.cell(row=row, column=8, value=float(item.total_cost))
            ws_machinery.cell(row=row, column=9, value=item.supplier or 'TBD')
            
            # Format currency
            for col in [5, 6, 7, 8]:
                ws_machinery.cell(row=row, column=col).number_format = '£#,##0.00'
    
    # Auto-adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()